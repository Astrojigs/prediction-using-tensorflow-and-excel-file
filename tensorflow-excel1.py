import tensorflow as tf
from tensorflow import keras
import numpy as np
import openpyxl as xl

workbook = xl.load_workbook('temperature data for neural networking_original.xlsx')

sheet = workbook['Sheet1']

# creating lists,
train_daysx = []
train_windx = []
train_tempx = []
train_rate = []  # good or bad
class_name = ['bad day', 'good day']

for row in range(2, 10):
    train_dayx_location = sheet.cell(row, 1)
    train_daysx.append(train_dayx_location.value)
    train_windx_location = sheet.cell(row, 2)
    train_windx.append(train_windx_location.value)
    train_tempx_location = sheet.cell(row, 3)
    train_tempx.append(train_tempx_location.value)
    train_rate_location = sheet.cell(row, 4)
    train_rate.append(train_rate_location.value)
print(train_rate)
train_inputs = [train_daysx, train_windx, train_tempx]
train_inputs_array = np.array(train_inputs).T
final_train_inputs_array = train_inputs_array.reshape(train_inputs_array.shape[0], 1, 3)
train_outputs_array = np.array(train_rate).T
final_train_outputs_array = train_outputs_array.reshape(train_outputs_array.shape[0], 1, 1)

# list for test inputs
test_daysx = []
test_windx = []
test_tempx = []

for row in range(10, 18):
    test_dayx_location = sheet.cell(row, 1)
    test_daysx.append(int(test_dayx_location.value))
    test_windx_location = sheet.cell(row, 2)
    test_windx.append(test_windx_location.value)
    test_tempx_location = sheet.cell(row, 3)
    test_tempx.append(test_tempx_location.value)
test_inputs = [test_daysx, test_windx, test_tempx]
test_inputs_array = np.array(test_inputs).T
final_test_inputs_array = test_inputs_array.reshape(test_inputs_array.shape[0], 1, 3)
print(final_train_inputs_array.shape)
print(final_train_outputs_array.shape)
print(final_test_inputs_array.shape)
model = keras.Sequential([
    keras.layers.Dense(4, input_shape=(1, 3), activation='relu'),
    keras.layers.Dense(6, activation='relu'),
    keras.layers.Dense(1, activation='sigmoid')])  # first number in bracket decides the no. of columns of 'prediction'

model.compile(optimizer='adam', loss='binary_crossentropy', metrics=['accuracy'])

model.fit(final_train_inputs_array, final_train_outputs_array, epochs=1000)
prediction = model.predict(final_test_inputs_array)

print(np.squeeze(prediction))
prediction_list = np.ndarray.tolist(np.squeeze(prediction))  # for converting predictions into a list
print(prediction_list)

s_row = 10  #starting row to store predictions
column_number = 5  # to store predicitions
# putting predictions in xlsx again
for element in prediction_list:
    print(element)
    element_position = sheet.cell(s_row, 5)
    element_position.value = class_name[round(element)]
    s_row+=1
    print('saved')
workbook.save('temperature data for neural networking_original_modified.xlsx')